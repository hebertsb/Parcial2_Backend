# sales/excel_exporter.py
"""
Exportador de reportes a formato Excel con estilos profesionales.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime


class ExcelExporter:
    """
    Clase para exportar reportes a formato Excel con formato profesional.
    """
    
    # Colores del tema
    COLOR_HEADER = "1A222E"  # Azul oscuro
    COLOR_SUBTITLE = "4A5568"  # Gris
    COLOR_TOTAL = "E2E8F0"  # Gris claro
    
    def __init__(self, report_data):
        self.report_data = report_data
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Reporte"
        self.current_row = 1
    
    def generate(self):
        """
        Genera el archivo Excel y retorna un BytesIO.
        """
        self._write_title()
        self._write_subtitle()
        self._write_headers()
        self._write_data()
        self._write_totals()
        self._apply_column_widths()
        
        # Guardar en memoria
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        return output
    
    def _write_title(self):
        """
        Escribe el título del reporte.
        """
        cell = self.worksheet.cell(row=self.current_row, column=1)
        cell.value = self.report_data.get('title', 'Reporte')
        cell.font = Font(size=16, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge cells para el título
        num_columns = len(self.report_data.get('headers', []))
        if num_columns > 1:
            self.worksheet.merge_cells(
                start_row=self.current_row,
                start_column=1,
                end_row=self.current_row,
                end_column=num_columns
            )
        
        self.current_row += 1
    
    def _write_subtitle(self):
        """
        Escribe el subtítulo del reporte.
        """
        subtitle = self.report_data.get('subtitle', '')
        if subtitle:
            cell = self.worksheet.cell(row=self.current_row, column=1)
            cell.value = subtitle
            cell.font = Font(size=11, italic=True, color="666666")
            cell.alignment = Alignment(horizontal='center')
            
            # Merge cells para el subtítulo
            num_columns = len(self.report_data.get('headers', []))
            if num_columns > 1:
                self.worksheet.merge_cells(
                    start_row=self.current_row,
                    start_column=1,
                    end_row=self.current_row,
                    end_column=num_columns
                )
            
            self.current_row += 1
        
        # Espacio en blanco
        self.current_row += 1
    
    def _write_headers(self):
        """
        Escribe los encabezados de las columnas.
        """
        headers = self.report_data.get('headers', [])
        
        for col_num, header in enumerate(headers, start=1):
            cell = self.worksheet.cell(row=self.current_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self._get_border()
        
        self.current_row += 1
    
    def _write_data(self):
        """
        Escribe los datos del reporte.
        """
        rows = self.report_data.get('rows', [])
        
        for row_data in rows:
            for col_num, value in enumerate(row_data, start=1):
                cell = self.worksheet.cell(row=self.current_row, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal='left' if col_num == 1 else 'center', vertical='center')
                cell.border = self._get_border(thin=True)
                
                # Aplicar formato de número si es necesario
                if isinstance(value, (int, float)) and col_num > 1:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            
            self.current_row += 1
    
    def _write_totals(self):
        """
        Escribe la fila de totales si existe.
        """
        totals = self.report_data.get('totals', {})
        
        if totals:
            # Espacio en blanco
            self.current_row += 1
            
            # Escribir cada total
            for key, value in totals.items():
                cell_label = self.worksheet.cell(row=self.current_row, column=1)
                cell_value = self.worksheet.cell(row=self.current_row, column=2)
                
                # Formatear el label
                label = key.replace('_', ' ').title()
                cell_label.value = label + ":"
                cell_label.font = Font(bold=True)
                cell_label.fill = PatternFill(start_color=self.COLOR_TOTAL, end_color=self.COLOR_TOTAL, fill_type="solid")
                
                # Valor
                cell_value.value = value
                cell_value.font = Font(bold=True)
                cell_value.fill = PatternFill(start_color=self.COLOR_TOTAL, end_color=self.COLOR_TOTAL, fill_type="solid")
                cell_value.alignment = Alignment(horizontal='right')
                
                self.current_row += 1
            
            # Añadir metadatos
            self.current_row += 1
            cell = self.worksheet.cell(row=self.current_row, column=1)
            cell.value = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            cell.font = Font(size=9, italic=True, color="999999")
    
    def _apply_column_widths(self):
        """
        Ajusta automáticamente el ancho de las columnas.
        """
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Máximo 50
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _get_border(self, thin=False):
        """
        Retorna un objeto Border para las celdas.
        """
        side_style = 'thin' if thin else 'medium'
        side = Side(style=side_style, color="000000")
        return Border(left=side, right=side, top=side, bottom=side)


def export_to_excel(report_data):
    """
    Función helper para exportar un reporte a Excel.
    
    Args:
        report_data (dict): Datos del reporte
    
    Returns:
        BytesIO: Archivo Excel en memoria
    """
    exporter = ExcelExporter(report_data)
    return exporter.generate()
