# sales/views_audit_reports.py
"""
Vistas para generar reportes de auditoría con exportación a PDF y Excel.
"""

from rest_framework import views, status
from rest_framework.response import Response
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime
from api.permissions import IsAdminUser

from .audit_report_generator import AuditReportGenerator, AuditSessionReportGenerator
from .excel_exporter import export_to_excel


class GenerateAuditReportView(views.APIView):
    """
    POST /api/sales/audit/generate-report/

    Genera reportes de bitácora con filtros dinámicos en PDF, Excel o JSON.

    Body:
    {
        "filters": {
            "user": "admin",                    // Filtrar por usuario
            "action_type": "AUTH",              // AUTH, CREATE, READ, UPDATE, DELETE, REPORT, PAYMENT, ML
            "start_date": "2025-01-01",         // Fecha inicio
            "end_date": "2025-10-21",           // Fecha fin
            "severity": "HIGH",                 // LOW, MEDIUM, HIGH, CRITICAL
            "success": true,                    // true/false
            "ip_address": "192.168.1.100",      // IP específica
            "endpoint": "/api/products/",       // Endpoint específico
            "limit": 1000                       // Límite de registros (default: 1000)
        },
        "format": "pdf"                         // pdf, excel, json
    }

    Ejemplos:
    1. Todas las acciones de un usuario en PDF:
       {
           "filters": {"user": "admin"},
           "format": "pdf"
       }

    2. Todos los errores del último mes en Excel:
       {
           "filters": {
               "success": false,
               "start_date": "2025-09-21",
               "end_date": "2025-10-21"
           },
           "format": "excel"
       }

    3. Todas las autenticaciones fallidas:
       {
           "filters": {
               "action_type": "AUTH",
               "success": false
           },
           "format": "pdf"
       }
    """
    permission_classes = [IsAdminUser]

    def post(self, request):
        filters = request.data.get('filters', {})
        output_format = request.data.get('format', 'json').lower()

        try:
            # Generar el reporte
            generator = AuditReportGenerator(filters)
            report_data = generator.generate()

            # Formatear según el formato solicitado
            if output_format == 'pdf':
                return self._export_to_pdf(report_data, filters)
            elif output_format == 'excel':
                return self._export_to_excel(report_data, filters)
            else:
                return Response({
                    'success': True,
                    'format': 'json',
                    'data': report_data
                }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _export_to_pdf(self, report_data, filters):
        """
        Exporta el reporte a PDF.
        """
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle
        from reportlab.lib.units import inch

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="bitacora_auditoria.pdf"'

        # Usar landscape para más columnas
        p = canvas.Canvas(response, pagesize=landscape(letter))
        width, height = landscape(letter)

        # Título
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, height - 50, report_data['title'])

        # Subtítulo
        p.setFont("Helvetica", 10)
        p.drawString(50, height - 70, report_data['subtitle'])

        # Fecha de generación
        p.setFont("Helvetica-Oblique", 9)
        p.drawString(50, height - 85, f"Generado: {report_data['metadata']['generado_en']}")

        y_position = height - 110

        # Resumen estadístico
        p.setFont("Helvetica-Bold", 11)
        p.drawString(50, y_position, "Resumen:")
        y_position -= 15

        p.setFont("Helvetica", 9)
        totals = report_data['totals']
        p.drawString(50, y_position, f"Total de registros: {totals['total_registros']}")
        p.drawString(200, y_position, f"Éxitos: {totals['total_exitos']}")
        p.drawString(300, y_position, f"Errores: {totals['total_errores']}")
        p.drawString(400, y_position, f"Tasa de error: {totals['tasa_error']}")
        y_position -= 25

        # Tabla de logs
        headers = report_data['headers']
        rows = report_data['rows'][:30]  # Limitar a 30 para que quepa en el PDF

        if rows:
            table_data = [headers] + rows

            # Calcular anchos de columna
            col_widths = [1.1*inch, 0.8*inch, 2*inch, 1.5*inch, 1*inch, 0.7*inch, 0.6*inch, 0.7*inch]

            table = Table(table_data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A222E')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))

            table_height = table.wrap(width, height)[1]
            table.drawOn(p, 50, y_position - table_height - 10)

            y_position -= table_height + 30

        # Pie de página
        p.setFont("Helvetica-Oblique", 8)
        p.drawString(50, 30, f"Sistema de Auditoría - {timezone.now().year}")
        p.drawRightString(width - 50, 30, f"Página 1 de 1")

        p.showPage()
        p.save()

        return response

    def _export_to_excel(self, report_data, filters):
        """
        Exporta el reporte a Excel.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from io import BytesIO

            wb = Workbook()
            ws = wb.active
            ws.title = "Bitácora"

            # Título
            ws['A1'] = report_data['title']
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells('A1:H1')

            # Subtítulo
            ws['A2'] = report_data['subtitle']
            ws['A2'].font = Font(size=10, italic=True)
            ws.merge_cells('A2:H2')

            # Fecha de generación
            ws['A3'] = f"Generado: {report_data['metadata']['generado_en']}"
            ws['A3'].font = Font(size=9, italic=True)
            ws.merge_cells('A3:H3')

            # Resumen
            row = 5
            ws[f'A{row}'] = 'RESUMEN'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

            totals = report_data['totals']
            ws[f'A{row}'] = f"Total de registros: {totals['total_registros']}"
            ws[f'C{row}'] = f"Éxitos: {totals['total_exitos']}"
            ws[f'E{row}'] = f"Errores: {totals['total_errores']}"
            ws[f'G{row}'] = f"Tasa de error: {totals['tasa_error']}"
            row += 2

            # Headers
            headers = report_data['headers']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1A222E", end_color="1A222E", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            row += 1

            # Datos
            for row_data in report_data['rows']:
                for col, value in enumerate(row_data, start=1):
                    ws.cell(row=row, column=col, value=str(value))
                row += 1

            # Ajustar anchos de columna
            column_widths = [20, 15, 35, 25, 15, 12, 12, 12]
            for i, width in enumerate(column_widths, start=1):
                ws.column_dimensions[chr(64 + i)].width = width

            # Guardar en memoria
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="bitacora_auditoria.xlsx"'

            return response

        except ImportError:
            return Response({
                'error': 'openpyxl no está instalado. Ejecuta: pip install openpyxl'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class GenerateSessionReportView(views.APIView):
    """
    POST /api/sales/audit/generate-session-report/

    Genera reportes de sesiones de usuarios con filtros.

    Body:
    {
        "filters": {
            "user": "admin",           // Filtrar por usuario
            "is_active": true,         // Solo sesiones activas/cerradas
            "start_date": "2025-01-01",
            "end_date": "2025-10-21",
            "limit": 500
        },
        "format": "pdf"                // pdf, excel, json
    }
    """
    permission_classes = [IsAdminUser]

    def post(self, request):
        filters = request.data.get('filters', {})
        output_format = request.data.get('format', 'json').lower()

        try:
            generator = AuditSessionReportGenerator(filters)
            report_data = generator.generate()

            if output_format == 'pdf':
                return self._export_to_pdf(report_data)
            elif output_format == 'excel':
                return self._export_to_excel(report_data)
            else:
                return Response({
                    'success': True,
                    'format': 'json',
                    'data': report_data
                }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _export_to_pdf(self, report_data):
        """Exporta reporte de sesiones a PDF."""
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle
        from reportlab.lib.units import inch

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="sesiones_usuarios.pdf"'

        p = canvas.Canvas(response, pagesize=letter)
        width, height = letter

        # Título
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, height - 50, report_data['title'])

        # Subtítulo
        p.setFont("Helvetica", 10)
        p.drawString(50, height - 70, report_data['subtitle'])

        y_position = height - 100

        # Resumen
        p.setFont("Helvetica-Bold", 11)
        p.drawString(50, y_position, "Resumen:")
        y_position -= 15

        p.setFont("Helvetica", 9)
        totals = report_data['totals']
        p.drawString(50, y_position, f"Total de sesiones: {totals['total_sesiones']}")
        p.drawString(200, y_position, f"Activas: {totals['sesiones_activas']}")
        p.drawString(300, y_position, f"Cerradas: {totals['sesiones_cerradas']}")
        p.drawString(420, y_position, f"Duración promedio: {totals['duracion_promedio_min']} min")
        y_position -= 25

        # Tabla
        headers = report_data['headers']
        rows = report_data['rows'][:25]

        if rows:
            table_data = [headers] + rows
            col_widths = [1.2*inch, 1.2*inch, 1.3*inch, 1.3*inch, 1.3*inch, 0.8*inch, 0.8*inch]

            table = Table(table_data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A222E')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))

            table_height = table.wrap(width, height)[1]
            table.drawOn(p, 50, y_position - table_height - 10)

        p.showPage()
        p.save()

        return response

    def _export_to_excel(self, report_data):
        """Exporta reporte de sesiones a Excel."""
        output = export_to_excel(report_data)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="sesiones_usuarios.xlsx"'

        return response
