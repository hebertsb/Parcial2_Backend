import time
import logging
import io
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from typing import Any, Dict, cast

from .models import VoiceCommand, VoiceCommandHistory
from .serializers import (
    VoiceCommandSerializer,
    VoiceCommandTextSerializer
)
from .voice_processor import VoiceCommandProcessor
import threading
from django.db import close_old_connections
from .handlers import handle_search_products, handle_recommend_products, handle_add_to_cart

logger = logging.getLogger(__name__)


class VoiceCommandViewSet(viewsets.ModelViewSet):
    """
    ViewSet para comandos de texto inteligentes
    
    Endpoints:
    - GET /text-commands/ - Listar comandos del usuario
    - GET /text-commands/{id}/ - Detalle de un comando
    - POST /text-commands/process/ - Procesar un comando de texto
    - GET /text-commands/history/ - Historial de todos los comandos
    - GET /text-commands/capabilities/ - Capacidades del sistema
    """
    
    serializer_class = VoiceCommandSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Solo comandos del usuario actual"""
        return VoiceCommand.objects.filter(user=self.request.user).prefetch_related('history')
    
    @action(detail=False, methods=['POST'], url_path='process')
    def process_text(self, request):
        """
        Procesa un comando de texto inteligente y genera el reporte correspondiente
        
        **Request:**
        ```json
        {
            "text": "generar reporte de ventas del último mes en PDF"
        }
        ```
        
        **Ejemplos de comandos válidos:**
        - "reporte de ventas del último mes"
        - "productos más vendidos esta semana"
        - "dashboard ejecutivo del mes de octubre"
        - "predicciones de ventas para los próximos 7 días"
        - "análisis RFM de clientes en Excel"
        - "ventas por cliente del año 2024"
        
        **Response:**
        ```json
        {
            "success": true,
            "data": {
                "id": 1,
                "transcribed_text": "generar reporte de ventas del último mes",
                "status": "EXECUTED",
                "command_type": "reporte",
                "result_data": {...},
                "processing_time_ms": 850
            },
            "message": "Comando procesado exitosamente"
        }
        ```
        """
        
        start_time = time.time()
        
        # Validar el input
        serializer = VoiceCommandTextSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                'success': False,
                'error': 'Datos inválidos',
                'details': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Evitar subscript sobre un posible None (Pylance puede marcar validated_data como Optional)
        # Forzar tipo a dict para que Pylance reconozca métodos como .get
        validated = cast(Dict[str, Any], (getattr(serializer, 'validated_data', None) or {}))
        text = validated.get('text', '')
        
        # Crear el comando en BD
        voice_command = VoiceCommand.objects.create(
            user=request.user,
            command_text=text,
            status='PROCESSING'
        )
        # Ayuda para Pylance: evitar warnings de atributo implícito de Django (id, fields)
        # Algunos analizadores estáticos no reconocen atributos creados por Django en tiempo de ejecución,
        # así que casteamos a Any para silenciar esos avisos en accesos como `voice_command.id`.
        voice_command = cast(Any, voice_command)
        
        VoiceCommandHistory.objects.create(
            voice_command=voice_command,
            stage='TEXT_INPUT',
            message=f'Comando recibido: "{text}"',
            data={'source': 'text_api'}
        )
        
        try:
            # Procesamiento en hilo con intento síncrono corto.
            # Si el procesamiento termina dentro del timeout lo devolvemos inmediatamente (auto-download posible).
            # Si no termina, devolvemos rápida/primer respuesta y dejamos el worker terminar en background.
            processor = VoiceCommandProcessor(user=request.user)

            # Worker que ejecuta el procesamiento y actualiza la BD
            def _worker_process(vc_id, txt):
                try:
                    # Asegurar conexiones limpias en el hilo
                    close_old_connections()
                    proc = VoiceCommandProcessor(user=request.user)
                    result = proc.process_command(txt)

                    vc = VoiceCommand.objects.get(id=vc_id)
                    vc = cast(Any, vc)
                    vc.command_type = result.get('command_type')
                    vc.interpreted_params = result.get('params', {})
                    vc.confidence_score = result.get('confidence', 0.0)

                    if result.get('success'):
                        vc.status = 'EXECUTED'
                        vc.result_data = result.get('result', {})
                        VoiceCommandHistory.objects.create(
                            voice_command=vc,
                            stage='EXECUTION_SUCCESS',
                            message='Comando ejecutado exitosamente',
                            data={'command_type': result.get('command_type')}
                        )
                    else:
                        vc.status = 'FAILED'
                        vc.error_message = result.get('error', 'Error desconocido')
                        VoiceCommandHistory.objects.create(
                            voice_command=vc,
                            stage='EXECUTION_FAILED',
                            message=f'Error: {vc.error_message}',
                            data={}
                        )

                    end_t = time.time()
                    vc.processing_time_ms = int((end_t - start_time) * 1000)
                    vc.save()
                except Exception as ex:
                    logger.error(f"❌ Error en worker de procesamiento: {ex}", exc_info=True)
                    try:
                        vc = VoiceCommand.objects.get(id=vc_id)
                        vc = cast(Any, vc)
                        vc.status = 'FAILED'
                        vc.error_message = str(ex)
                        vc.save()
                    except Exception:
                        pass
                finally:
                    # Cerrar conexión del hilo
                    close_old_connections()

            # Capturar id en una variable simple para evitar accesos directos que Pylance marque
            vc_id_local = getattr(voice_command, 'id')

            # Iniciar worker en hilo
            worker = threading.Thread(target=_worker_process, args=(vc_id_local, text), daemon=True)
            worker.start()

            # Intentar esperar un tiempo corto para respuesta rápida (umbral en segundos)
            QUICK_TIMEOUT = 1.0  # seconds; ajustar si se desea más/menos tolerancia
            worker.join(timeout=QUICK_TIMEOUT)

            if worker.is_alive():
                # Todavía procesando: respondemos rápidamente indicando que está en background
                logger.info(f"⏳ Comando {vc_id_local} en procesamiento asíncrono (took > {QUICK_TIMEOUT}s)")
                return Response({
                    'success': True,
                    'data': {
                        'id': vc_id_local,
                        'status': 'PROCESSING',
                        'message': 'Procesamiento en segundo plano. Puede descargar cuando el comando esté EXECUTED.'
                    }
                }, status=status.HTTP_202_ACCEPTED)
            else:
                # Worker terminó rápido: cargar objeto actualizado y devolver resultado completo
                voice_command.refresh_from_db()
                serializer = VoiceCommandSerializer(voice_command)
                return Response({
                    'success': True,
                    'data': serializer.data,
                    'message': 'Comando procesado exitosamente'
                }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"❌ Error al iniciar procesamiento: {e}", exc_info=True)
            voice_command.status = 'FAILED'
            voice_command.error_message = str(e)
            voice_command.save()
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['POST'], url_path='chat/process')
    def chat_process(self, request):
        """Endpoint conversacional para búsqueda, recomendaciones y añadir al carrito.

        Request JSON: {"text": "...", "intent_hint": "add_to_cart|search|recommend"}
        """
        serializer = VoiceCommandTextSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({'success': False, 'error': 'Datos inválidos', 'details': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

        validated = cast(Dict[str, Any], (getattr(serializer, 'validated_data', None) or {}))
        text = validated.get('text', '').strip()
        intent_hint = validated.get('intent_hint')

        # Guardar el comando en BD para trazabilidad
        vc = VoiceCommand.objects.create(user=request.user, command_text=text, status='PROCESSING')
        VoiceCommandHistory.objects.create(voice_command=vc, stage='TEXT_INPUT', message=f'Chat input: {text}', data={})

        # Usar IA (VoiceCommandProcessor) siempre para interpretar la entrada del usuario.
        # Esto permite manejar preguntas libres como "¿tienen heladera LG?" o "qué me recomiendas"
        t_low = text.lower()
        intent = intent_hint
        processor = VoiceCommandProcessor(user=request.user)
        parsed = processor.process_command(text)

        # Si el parser identifica un tipo de comando ecommerce claro, lo usamos.
        if not intent and parsed:
            p_cmd = parsed.get('command_type')
            if p_cmd in ('recommend', 'search', 'add_to_cart'):
                intent = p_cmd

        # Si todavía no está claro, aplicar heurísticas más simples
        if not intent:
            if any(w in t_low for w in ['recomienda', 'recomiéndame', 'sugiere', 'sugerir']):
                intent = 'recommend'
            elif any(k in t_low for k in ['añade', 'agrega', 'al carrito', 'alcarrito', 'añadir', 'agregar', 'pon en el carrito', 'comprame', 'comprarme']):
                intent = 'add_to_cart'
            elif any(w in t_low for w in ['buscar', 'tienes', 'hay', 'muéstrame', 'mostrar', 'mostrarme']):
                intent = 'search'
            else:
                # Por defecto usar 'search' para preguntas abiertas sobre productos
                intent = 'search'
                # si parser sugiere un reporte detectamos y ofrecer formatos de descarga
                if parsed.get('command_type') in ('reporte', 'reporte_pdf', 'reporte_excel', 'report'):
                    # Si el parser entiende que es un reporte, revisar formato esperado
                    fmt = None
                    try:
                        fmt = parsed.get('params', {}).get('format') or parsed.get('format')
                    except Exception:
                        fmt = None

                    # Decidir comportamiento para reportes:
                    # - Si el usuario pidió explícitamente una descarga ("descarg" o menciona PDF/Excel/Word),
                    #   ofrecer solo formatos binarios permitidos.
                    # - No sugerir nunca "json" como opción de descarga.
                    allowed_bin = {'pdf', 'excel', 'xlsx', 'docx', 'word'}
                    user_requested_download = ('descarg' in t_low) or any(k in t_low for k in ['pdf', 'excel', 'word', 'docx', 'xlsx'])
                    fmt_l = str(fmt).lower() if fmt else None

                    if user_requested_download or (fmt_l and fmt_l in allowed_bin):
                        formats = ['pdf', 'xlsx', 'docx']
                        vc.status = 'EXECUTED'
                        vc.result_data = {'type': 'offer_download', 'report': parsed.get('report_name'), 'formats': formats}
                        vc.save()
                        return Response({
                            'success': True,
                            'intent': 'offer_download_formats',
                            'message': 'Puedo generar el reporte en varios formatos. ¿Cuál prefieres?',
                            'actions': [{'type': 'offer_download_formats', 'formats': formats}],
                            'report_hint': parsed.get('report_name')
                        })

                    # Si el parser pidió JSON o el usuario NO solicitó descarga, devolver un resumen en texto
                    if fmt_l and fmt_l == 'json':
                        # Construir un resumen simple para mostrar en la UI
                        parsed_result = parsed.get('result') or {}
                        report_info = parsed_result.get('report_info', {}) if isinstance(parsed_result, dict) else {}
                        metadata = parsed_result.get('metadata', {}) if isinstance(parsed_result, dict) else {}
                        total = metadata.get('total_records', 'N/A')
                        report_name = report_info.get('name') or parsed.get('params', {}).get('report_type') or parsed.get('report_name')

                        # Intentar obtener una vista previa de la primera fila
                        preview = ''
                        try:
                            data_section = parsed_result.get('data') if isinstance(parsed_result, dict) else None
                            if isinstance(data_section, dict) and 'rows' in data_section and isinstance(data_section['rows'], list) and data_section['rows']:
                                first_row = data_section['rows'][0]
                                preview = str(first_row)
                            elif isinstance(data_section, list) and data_section:
                                preview = str(data_section[0])
                        except Exception:
                            preview = ''

                        summary_text = f"{report_name}: {total} registros." + (f" Ejemplo: {preview}" if preview else '')

                        vc.status = 'EXECUTED'
                        vc.result_data = {'type': 'report_summary', 'report': report_name, 'summary': summary_text, 'details': parsed_result}
                        vc.save()

                        return Response({
                            'success': True,
                            'intent': 'report_summary',
                            'message': 'Aquí tienes un resumen del reporte',
                            'summary': summary_text,
                            'details': parsed_result,
                            'report_hint': report_name
                        })

                # si el parser devuelve un intent ecommerce claro, usarlo
                intent = parsed.get('command_type') if parsed and parsed.get('command_type') in ('search', 'recommend', 'add_to_cart') else 'search'

        # Ejecutar intent
        try:
            if intent == 'search':
                res = handle_search_products(text)
                vc.status = 'EXECUTED'
                vc.result_data = {'type': 'search', 'result': res}
                vc.save()
                return Response({'success': True, 'intent': 'search', 'result': res})

            if intent == 'recommend' or intent == 'recommend_products':
                res = handle_recommend_products(request.user, text)
                vc.status = 'EXECUTED'
                vc.result_data = {'type': 'recommend', 'result': res}
                vc.save()
                return Response({'success': True, 'intent': 'recommend', 'result': res})

            if intent == 'add_to_cart':
                # intentar extraer id si existe
                res = handle_add_to_cart(request.user, text)
                if res.get('success'):
                    vc.status = 'EXECUTED'
                    vc.result_data = {'type': 'add_to_cart', 'result': res}
                    vc.save()
                    return Response({'success': True, 'intent': 'add_to_cart', 'result': res})
                else:
                    vc.status = 'FAILED'
                    vc.error_message = res.get('error')
                    vc.save()
                    return Response({'success': False, 'intent': 'add_to_cart', 'error': res.get('error')}, status=status.HTTP_400_BAD_REQUEST)

            # Default
            return Response({'success': False, 'error': 'Intent no soportado'}, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            logger.error(f"Error al procesar chat intent: {e}", exc_info=True)
            vc.status = 'FAILED'
            vc.error_message = str(e)
            vc.save()
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['GET'], url_path='history')
    def history(self, request):
        """
        Obtiene el historial de todos los comandos del usuario
        """
        
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        
        return Response({
            'success': True,
            'count': queryset.count(),
            'data': serializer.data
        })
    
    @action(detail=False, methods=['GET'], url_path='capabilities')
    def capabilities(self, request):
        """
        Retorna las capacidades del sistema de comandos inteligentes
        """
        
        return Response({
            'success': True,
            'data': {
                'system_name': 'Sistema de Comandos Inteligentes',
                'version': '2.0',
                'supported_commands': [
                    'Reportes de Ventas',
                    'Reportes de Productos',
                    'Reportes de Clientes',
                    'Análisis Avanzados (RFM, ABC)',
                    'Predicciones ML',
                    'Dashboard Ejecutivo'
                ],
                'examples': [
                    'reporte de ventas del último mes',
                    'productos más vendidos esta semana',
                    'dashboard ejecutivo de octubre',
                    'predicciones de ventas para 7 días',
                    'análisis RFM de clientes en Excel',
                    'ventas por cliente del año 2024',
                    'análisis ABC de productos en PDF'
                ],
                'supported_formats': ['PDF', 'Excel', 'Word'],
                'supported_date_ranges': [
                    'hoy', 'ayer', 'esta semana', 'este mes', 'este año',
                    'última semana', 'último mes', 'último año',
                    'últimos X días',
                    'mes de [nombre]',
                    'año [número]',
                    'del DD/MM/YYYY al DD/MM/YYYY'
                ]
            }
        })
    
    @action(detail=True, methods=['GET'], url_path='download/pdf')
    def download_pdf(self, request, pk=None):
        """
        Descarga el reporte en formato PDF
        
        GET /api/voice-commands/{id}/download/pdf/
        """
        try:
            voice_command = self.get_object()
            
            # Verificar que el comando se ejecutó exitosamente
            if voice_command.status != 'EXECUTED':
                return Response({
                    'error': 'El comando no se ejecutó correctamente. No se puede generar PDF.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Importar librerías de PDF
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            
            # Crear buffer
            buffer = io.BytesIO()
            
            # Obtener datos del reporte
            result_data = voice_command.result_data or {}
            report_info = result_data.get('report_info', {})
            parameters = result_data.get('parameters', {})
            metadata = result_data.get('metadata', {})
            
            # Crear documento PDF con márgenes balanceados
            doc = SimpleDocTemplate(
                buffer,
                pagesize=letter,
                leftMargin=36,
                rightMargin=36,
                topMargin=36,
                bottomMargin=36,
            )
            elements = []
            styles = getSampleStyleSheet()
            
            # Estilo personalizado para título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1a73e8'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            
            # Título
            title = Paragraph(report_info.get('name', 'Reporte de Comandos Inteligentes'), title_style)
            elements.append(title)
            elements.append(Spacer(1, 0.3*inch))
            
            # Información del comando
            command_info = [
                ['Comando Original:', voice_command.command_text],
                ['Estado:', voice_command.status],
                ['Confianza:', f"{(voice_command.confidence_score * 100):.0f}%" if voice_command.confidence_score else "N/A"],
                ['Tiempo de Procesamiento:', f"{voice_command.processing_time_ms}ms" if voice_command.processing_time_ms else "N/A"],
                ['Generado por:', voice_command.user.username],
                ['Fecha de Generación:', voice_command.created_at.strftime('%d/%m/%Y %H:%M:%S')],
            ]
            
            # Centrar la tabla de información del comando y ajustar anchos relativos
            try:
                t1_col_widths = [doc.width * 0.35, doc.width * 0.65]
            except Exception:
                t1_col_widths = [2.5 * inch, 4 * inch]

            t1 = Table(command_info, colWidths=t1_col_widths, hAlign='CENTER')
            t1.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8f0fe')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            elements.append(t1)
            elements.append(Spacer(1, 0.5*inch))
            
            # Parámetros del reporte
            # Añadir resumen de filtros aplicados justo debajo del título (si hay)
            if parameters:
                # Título pequeño para filtros
                filtros_title = Paragraph('<b>Filtros aplicados</b>', styles['Heading3'])
                elements.append(filtros_title)
                elements.append(Spacer(1, 0.08*inch))

                # Construir lista de pares (clave, valor)
                filter_rows = []
                date_range = parameters.get('date_range', {})
                if date_range.get('description'):
                    filter_rows.append(['Período', date_range['description']])
                if date_range.get('start'):
                    filter_rows.append(['Fecha Inicio', date_range['start'][:10]])
                if date_range.get('end'):
                    filter_rows.append(['Fecha Fin', date_range['end'][:10]])

                # Otros parámetros planos
                for k, v in parameters.items():
                    if k == 'date_range':
                        continue
                    # Mostrar arrays y dicts de forma legible
                    if isinstance(v, (list, tuple)):
                        val = ', '.join(str(x) for x in v)
                    elif isinstance(v, dict):
                        val = ', '.join(f"{kk}:{vv}" for kk, vv in v.items())
                    else:
                        val = str(v)
                    filter_rows.append([str(k).replace('_', ' ').title(), val])

                if filter_rows:
                    try:
                        fr_col_widths = [doc.width * 0.3, doc.width * 0.7]
                    except Exception:
                        fr_col_widths = [2 * inch, 4 * inch]

                    ft = Table(filter_rows, colWidths=fr_col_widths, hAlign='CENTER')
                    ft.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f1f3f4')),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 9),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey)
                    ]))
                    elements.append(ft)
                    elements.append(Spacer(1, 0.18*inch))
            
            # Descripción del reporte
            if report_info.get('description'):
                desc_title = Paragraph('<b>Descripción</b>', styles['Heading2'])
                elements.append(desc_title)
                elements.append(Spacer(1, 0.1*inch))
                desc = Paragraph(report_info['description'], styles['BodyText'])
                elements.append(desc)
                elements.append(Spacer(1, 0.3*inch))

            # ============ DATOS DEL REPORTE (LA TABLA PRINCIPAL) ============
            # Esta es la parte que faltaba: escribir los datos reales del reporte
            report_data = result_data.get('data', {})

            if report_data:
                # Título de los datos
                data_title = Paragraph('<b>Datos del Reporte</b>', styles['Heading2'])
                elements.append(data_title)
                elements.append(Spacer(1, 0.2*inch))

                # Construir tabla de datos
                # Intentar extraer headers y rows del reporte
                headers = report_data.get('headers', [])
                rows = report_data.get('rows', [])

                # Si no hay estructura headers/rows, intentar otras estructuras
                if not headers and not rows:
                    # ESTRUCTURA ML: Predicciones con 'predictions'
                    if 'predictions' in report_data:
                        predictions = report_data['predictions']
                        if len(predictions) > 0:
                            # Headers para predicciones ML
                            headers = ['Fecha', 'Ventas Predichas', 'Límite Inferior', 'Límite Superior', 'Confianza']
                            rows = []
                            for pred in predictions[:50]:  # Limitar a 50
                                rows.append([
                                    pred.get('date', 'N/A'),
                                    f"${pred.get('predicted_sales', 0):,.2f}",
                                    f"${pred.get('lower_bound', 0):,.2f}",
                                    f"${pred.get('upper_bound', 0):,.2f}",
                                    f"{pred.get('confidence', 0) * 100:.0f}%"
                                ])
                    # Estructura alternativa: lista de diccionarios
                    elif isinstance(report_data, list) and len(report_data) > 0:
                        if isinstance(report_data[0], dict):
                            headers = list(report_data[0].keys())
                            rows = [[str(item.get(h, '')) for h in headers] for item in report_data]
                    # Estructura alternativa: diccionario con 'data'
                    elif 'data' in report_data and isinstance(report_data['data'], list):
                        data_list = report_data['data']
                        if len(data_list) > 0 and isinstance(data_list[0], dict):
                            headers = list(data_list[0].keys())
                            rows = [[str(item.get(h, '')) for h in headers] for item in data_list]

                if headers and rows:
                    # Preparar datos para la tabla (headers + rows)
                    table_data = [headers] + rows[:50]  # Limitar a 50 filas para evitar PDFs muy grandes

                    # Calcular anchos de columna para ocupar el ancho utilizable del documento
                    try:
                        col_count = len(headers)
                        col_widths = [doc.width / col_count for _ in range(col_count)]
                    except Exception:
                        col_widths = None

                    # Crear tabla centrada
                    data_table = Table(table_data, colWidths=col_widths, hAlign='CENTER')
                    data_table.setStyle(TableStyle([
                        # Estilo del header
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a73e8')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        # Estilo de las filas
                        ('FONTSIZE', (0, 1), (-1, -1), 8),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f3f4')]),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ]))
                    elements.append(data_table)

                    # Nota si se truncaron filas
                    if len(rows) > 50:
                        note = Paragraph(
                            f'<i>Mostrando 50 de {len(rows)} registros. Descargue el reporte completo en Excel.</i>',
                            styles['Normal']
                        )
                        elements.append(Spacer(1, 0.1*inch))
                        elements.append(note)

                    elements.append(Spacer(1, 0.3*inch))

                    # Si es una predicción ML, agregar el resumen
                    if 'summary' in report_data:
                        summary = report_data['summary']
                        summary_title = Paragraph('<b>Resumen de Predicciones</b>', styles['Heading2'])
                        elements.append(summary_title)
                        elements.append(Spacer(1, 0.2*inch))

                        summary_data = [
                            ['Total de días predichos:', str(summary.get('total_days', 'N/A'))],
                            ['Ventas totales predichas:', f"${summary.get('total_predicted_sales', 0):,.2f}"],
                            ['Promedio diario predicho:', f"${summary.get('average_daily_sales', 0):,.2f}"],
                            ['Promedio histórico:', f"${summary.get('historical_average', 0):,.2f}"],
                            ['Tasa de crecimiento:', f"{summary.get('growth_rate_percent', 0):+.2f}%"],
                        ]

                        # Centrar la tabla de resumen y ajustar anchos relativos
                        try:
                            summary_col_widths = [doc.width * 0.6, doc.width * 0.4]
                        except Exception:
                            summary_col_widths = [3 * inch, 2 * inch]

                        summary_table = Table(summary_data, colWidths=summary_col_widths, hAlign='CENTER')
                        summary_table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8f0fe')),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, -1), 10),
                            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
                        ]))
                        elements.append(summary_table)
                        elements.append(Spacer(1, 0.3*inch))

            # Nota al pie
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.grey,
                alignment=TA_CENTER
            )
            footer = Paragraph(
                f'Generado por Sistema de Comandos Inteligentes v2.0 | {metadata.get("generated_at", "")[:10]}',
                footer_style
            )
            elements.append(Spacer(1, 0.5*inch))
            elements.append(footer)
            
            # Construir PDF
            doc.build(elements)
            
            # Preparar respuesta
            buffer.seek(0)
            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            filename = f'reporte_{voice_command.id}_{voice_command.created_at.strftime("%Y%m%d")}.pdf'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            logger.info(f"📄 PDF generado exitosamente para comando {voice_command.id}")
            return response
            
        except Exception as e:
            logger.error(f"❌ Error al generar PDF: {e}", exc_info=True)
            return Response({
                'error': f'Error al generar PDF: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['GET'], url_path='download/excel')
    def download_excel(self, request, pk=None):
        """
        Descarga el reporte en formato Excel
        
        GET /api/voice-commands/{id}/download/excel/
        """
        try:
            voice_command = self.get_object()
            
            # Verificar que el comando se ejecutó exitosamente
            if voice_command.status != 'EXECUTED':
                return Response({
                    'error': 'El comando no se ejecutó correctamente. No se puede generar Excel.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Importar librerías de Excel
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            # Tipado para ayudar a Pylance
            from typing import cast
            try:
                from openpyxl.worksheet.worksheet import Worksheet
            except Exception:
                Worksheet = None  # pragma: no cover - typing fallback

            # Crear workbook
            wb = Workbook()
            # wb.active devuelve siempre una Worksheet en runtime, pero algunos analizadores lo tratan como Optional
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet(title="Reporte")
            else:
                # forzar tipo para el analizador
                if 'Worksheet' in globals() and Worksheet is not None:
                    ws = cast(Worksheet, ws)

            ws.title = "Reporte"
            
            # Obtener datos del reporte
            result_data = voice_command.result_data or {}
            report_info = result_data.get('report_info', {})
            parameters = result_data.get('parameters', {})
            
            # Estilos
            title_font = Font(name='Calibri', size=16, bold=True, color='1a73e8')
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1a73e8', end_color='1a73e8', fill_type='solid')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Título
            ws['A1'] = report_info.get('name', 'Reporte de Comandos Inteligentes')
            ws['A1'].font = title_font
            ws.merge_cells('A1:B1')

            # Pequeño resumen de filtros para aparecer como encabezado en el Excel
            if parameters:
                try:
                    parts = []
                    date_range = parameters.get('date_range', {})
                    if date_range.get('description'):
                        parts.append(date_range['description'])
                    if date_range.get('start') and date_range.get('end'):
                        parts.append(f"{date_range['start'][:10]} - {date_range['end'][:10]}")

                    for k, v in parameters.items():
                        if k == 'date_range':
                            continue
                        if v is None or v == '':
                            continue
                        if isinstance(v, (list, tuple)):
                            vv = ','.join(str(x) for x in v)
                        elif isinstance(v, dict):
                            vv = ','.join(f"{kk}:{vv}" for kk, vv in v.items())
                        else:
                            vv = str(v)
                        parts.append(f"{str(k).replace('_', ' ').title()}: {vv}")

                    summary_text = ' | '.join(parts)[:500]
                    ws.merge_cells('A2:B2')
                    ws['A2'] = summary_text
                    ws['A2'].font = Font(italic=True, size=10)
                    ws['A2'].alignment = Alignment(horizontal='center')
                except Exception:
                    # Si algo falla, no rompemos la generación del Excel
                    pass
            
            # Información del comando
            row = 3
            ws[f'A{row}'] = 'Comando Original:'
            ws[f'B{row}'] = voice_command.command_text
            ws[f'A{row}'].font = Font(bold=True)
            
            row += 1
            ws[f'A{row}'] = 'Estado:'
            ws[f'B{row}'] = voice_command.status
            ws[f'A{row}'].font = Font(bold=True)
            
            row += 1
            ws[f'A{row}'] = 'Confianza:'
            ws[f'B{row}'] = f"{(voice_command.confidence_score * 100):.0f}%" if voice_command.confidence_score else "N/A"
            ws[f'A{row}'].font = Font(bold=True)
            
            row += 1
            ws[f'A{row}'] = 'Tiempo de Procesamiento:'
            ws[f'B{row}'] = f"{voice_command.processing_time_ms}ms" if voice_command.processing_time_ms else "N/A"
            ws[f'A{row}'].font = Font(bold=True)
            
            row += 1
            ws[f'A{row}'] = 'Generado por:'
            ws[f'B{row}'] = voice_command.user.username
            ws[f'A{row}'].font = Font(bold=True)
            
            row += 1
            ws[f'A{row}'] = 'Fecha de Generación:'
            ws[f'B{row}'] = voice_command.created_at.strftime('%d/%m/%Y %H:%M:%S')
            ws[f'A{row}'].font = Font(bold=True)
            
            # Parámetros del reporte
            if parameters:
                row += 2
                ws[f'A{row}'] = 'Parámetros del Reporte'
                ws[f'A{row}'].font = Font(bold=True, size=12)
                ws.merge_cells(f'A{row}:B{row}')
                
                date_range = parameters.get('date_range', {})
                
                if date_range.get('description'):
                    row += 1
                    ws[f'A{row}'] = 'Período:'
                    ws[f'B{row}'] = date_range['description']
                    ws[f'A{row}'].font = Font(bold=True)
                
                if date_range.get('start'):
                    row += 1
                    ws[f'A{row}'] = 'Fecha Inicio:'
                    ws[f'B{row}'] = date_range['start'][:10]
                    ws[f'A{row}'].font = Font(bold=True)
                
                if date_range.get('end'):
                    row += 1
                    ws[f'A{row}'] = 'Fecha Fin:'
                    ws[f'B{row}'] = date_range['end'][:10]
                    ws[f'A{row}'].font = Font(bold=True)
                
                if parameters.get('group_by'):
                    row += 1
                    ws[f'A{row}'] = 'Agrupado por:'
                    ws[f'B{row}'] = parameters['group_by'].title()
                    ws[f'A{row}'].font = Font(bold=True)
                
                if parameters.get('limit'):
                    row += 1
                    ws[f'A{row}'] = 'Límite:'
                    ws[f'B{row}'] = parameters['limit']
                    ws[f'A{row}'].font = Font(bold=True)

            # ============ DATOS DEL REPORTE (LA TABLA PRINCIPAL) ============
            # Esta es la parte que faltaba: escribir los datos reales del reporte
            report_data = result_data.get('data', {})

            if report_data:
                row += 3  # Espacio
                ws[f'A{row}'] = 'Datos del Reporte'
                ws[f'A{row}'].font = Font(bold=True, size=14, color='1a73e8')
                ws.merge_cells(f'A{row}:E{row}')
                row += 2

                # Intentar extraer headers y rows del reporte
                headers = report_data.get('headers', [])
                rows = report_data.get('rows', [])

                # Si no hay estructura headers/rows, intentar otras estructuras
                if not headers and not rows:
                    # ESTRUCTURA ML: Predicciones con 'predictions'
                    if 'predictions' in report_data:
                        predictions = report_data['predictions']
                        if len(predictions) > 0:
                            # Headers para predicciones ML
                            headers = ['Fecha', 'Ventas Predichas', 'Límite Inferior', 'Límite Superior', 'Confianza']
                            rows = []
                            for pred in predictions:
                                rows.append([
                                    pred.get('date', 'N/A'),
                                    f"${pred.get('predicted_sales', 0):,.2f}",
                                    f"${pred.get('lower_bound', 0):,.2f}",
                                    f"${pred.get('upper_bound', 0):,.2f}",
                                    f"{pred.get('confidence', 0) * 100:.0f}%"
                                ])
                    # Estructura alternativa: lista de diccionarios
                    elif isinstance(report_data, list) and len(report_data) > 0:
                        if isinstance(report_data[0], dict):
                            headers = list(report_data[0].keys())
                            rows = [[item.get(h, '') for h in headers] for item in report_data]
                    # Estructura alternativa: diccionario con 'data'
                    elif 'data' in report_data and isinstance(report_data['data'], list):
                        data_list = report_data['data']
                        if len(data_list) > 0 and isinstance(data_list[0], dict):
                            headers = list(data_list[0].keys())
                            rows = [[item.get(h, '') for h in headers] for item in data_list]

                if headers and rows:
                    # Escribir encabezados
                    for col_idx, header in enumerate(headers, start=1):
                        cell = ws.cell(row=row, column=col_idx)
                        cell.value = header
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border

                    row += 1

                    # Escribir filas de datos
                    for row_data in rows:
                        for col_idx, value in enumerate(row_data, start=1):
                            cell = ws.cell(row=row, column=col_idx)
                            # Manejar valores que pueden ser strings con formato especial
                            cell.value = str(value) if value is not None else ''
                            cell.alignment = Alignment(horizontal='left' if col_idx == 1 else 'center', vertical='center')
                            cell.border = border
                        row += 1

                    # Ajustar ancho de columnas automáticamente para los datos
                    for col_idx in range(1, len(headers) + 1):
                        # Usar get_column_letter para evitar problemas con celdas fusionadas
                        column_letter = get_column_letter(col_idx)
                        max_length = len(str(headers[col_idx-1])) if col_idx <= len(headers) else 10

                        # Calcular ancho máximo basado en contenido
                        for row_data in rows[:100]:  # Revisar primeras 100 filas
                            if col_idx <= len(row_data):
                                cell_value = str(row_data[col_idx-1])
                                max_length = max(max_length, len(cell_value))

                        # Aplicar ancho (máximo 50 caracteres)
                        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

                    # Si es una predicción ML, agregar el resumen
                    if 'summary' in report_data:
                        summary = report_data['summary']
                        row += 3  # Espacio

                        # Título del resumen
                        ws[f'A{row}'] = 'Resumen de Predicciones'
                        ws[f'A{row}'].font = Font(bold=True, size=12, color='1a73e8')
                        ws.merge_cells(f'A{row}:B{row}')
                        row += 2

                        # Datos del resumen
                        summary_items = [
                            ('Total de días predichos:', str(summary.get('total_days', 'N/A'))),
                            ('Ventas totales predichas:', f"${summary.get('total_predicted_sales', 0):,.2f}"),
                            ('Promedio diario predicho:', f"${summary.get('average_daily_sales', 0):,.2f}"),
                            ('Promedio histórico:', f"${summary.get('historical_average', 0):,.2f}"),
                            ('Tasa de crecimiento:', f"{summary.get('growth_rate_percent', 0):+.2f}%"),
                        ]

                        for label, value in summary_items:
                            ws[f'A{row}'] = label
                            ws[f'A{row}'].font = Font(bold=True)
                            ws[f'A{row}'].fill = PatternFill(start_color='e8f0fe', end_color='e8f0fe', fill_type='solid')

                            ws[f'B{row}'] = value
                            ws[f'B{row}'].fill = PatternFill(start_color='e8f0fe', end_color='e8f0fe', fill_type='solid')

                            row += 1
            else:
                # Mantener anchos originales si no hay datos del reporte
                ws.column_dimensions['A'].width = 25
                ws.column_dimensions['B'].width = 50

            # Guardar en buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Preparar respuesta
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f'reporte_{voice_command.id}_{voice_command.created_at.strftime("%Y%m%d")}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            logger.info(f"📊 Excel generado exitosamente para comando {voice_command.id}")
            return response
            
        except Exception as e:
            logger.error(f"❌ Error al generar Excel: {e}", exc_info=True)
            return Response({
                'error': f'Error al generar Excel: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
